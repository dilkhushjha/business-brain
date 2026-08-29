from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .tally import TOTAL_MARKERS, clean_tally_text, normalize_tally_rows

TALLY_HEADERS = {
    "date", "voucher no", "voucher number", "voucher type", "party name",
    "particulars", "stock item", "item name", "quantity", "qty", "rate",
    "amount", "value", "taxable value", "cgst", "sgst", "igst", "gst",
}


def _header_score(values: list[Any]) -> int:
    score = 0
    for value in values:
        text = clean_tally_text(value)
        if text and text.lower() in TALLY_HEADERS:
            score += 2
        elif text:
            lowered = text.lower()
            if any(token in lowered for token in ("party", "voucher", "date", "item", "qty", "rate", "amount", "tax")):
                score += 1
    return score


def _find_header_row(ws, scan_limit: int = 30) -> int | None:
    best_row: int | None = None
    best_score = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, scan_limit)), start=1):
        score = _header_score(list(row))
        if score > best_score:
            best_score = score
            best_row = row_idx
    return best_row


def inspect_tally_workbook(path: Path) -> dict[str, Any]:
    """Inspect every worksheet and identify likely Tally data tables."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: list[dict[str, Any]] = []
    for ws in workbook.worksheets:
        header_row = _find_header_row(ws)
        headers: list[str] = []
        if header_row:
            headers = [clean_tally_text(v) or "" for v in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
        score = _header_score(headers)
        sheets.append({
            "name": ws.title,
            "rows": max(ws.max_row - (header_row or 0), 0),
            "columns": len([h for h in headers if h]),
            "header_row": header_row,
            "tally_score": score,
            "likely_data": score >= 3 and bool(header_row),
            "headers": [h for h in headers if h],
        })
    workbook.close()
    likely = [sheet for sheet in sheets if sheet["likely_data"]]
    return {
        "path": str(path),
        "sheet_count": len(sheets),
        "likely_data_sheets": len(likely),
        "sheets": sheets,
    }


def read_tally_sheet(path: Path, sheet_name: str, header_row: int | None = None) -> list[dict[str, Any]]:
    """Read one detected Tally sheet, skipping blank/report rows and normalizing values."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Sheet not found: {sheet_name}")
    ws = workbook[sheet_name]
    header_row = header_row or _find_header_row(ws)
    if not header_row:
        workbook.close()
        raise ValueError(f"Could not detect a header row in sheet: {sheet_name}")

    raw_headers = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    headers = [clean_tally_text(v) or f"column_{i + 1}" for i, v in enumerate(raw_headers)]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        rows.append(row)
    workbook.close()
    return normalize_tally_rows(rows)
